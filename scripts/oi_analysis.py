"""
oi_analysis.py
---------------------------------
Reusable functions for Options OI Analysis & Max Pain
using the ICICI Breeze API.

Usage (in a notebook or another script):
----------------------------------------
from scripts.oi_analysis import (
    fetch_option_chain_data,
    calculate_pcr_max_pain,
    append_to_excel,
    plot_open_interest,
    plot_total_loss,
    fetch_underlying_ltp
)

data = fetch_option_chain_data("NIFTY", "2025-09-25T06:00:00.000Z")
results = calculate_pcr_max_pain(data)
print(results)
"""

import os
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
from datetime import datetime

# Import Breeze client (local-only credentials in scripts/breeze_client.py)
from scripts.breeze_client import get_breeze

# Initialize Breeze session
breeze = get_breeze()

# Ensure data folder exists
os.makedirs("data", exist_ok=True)


# ----------------------------
# Fetch Option Chain Data
# ----------------------------
def fetch_option_chain_data(stock_code, expiry_date):
    """Fetch option chain data (calls + puts) for given stock and expiry."""
    try:
        call_response = breeze.get_option_chain_quotes(
            stock_code=stock_code, exchange_code="NFO",
            product_type="options", expiry_date=expiry_date, right="call"
        )
        put_response = breeze.get_option_chain_quotes(
            stock_code=stock_code, exchange_code="NFO",
            product_type="options", expiry_date=expiry_date, right="put"
        )

        call_data = pd.DataFrame(call_response.get("Success", []))
        put_data = pd.DataFrame(put_response.get("Success", []))

        if not call_data.empty:
            call_data["Option_Type"] = "Call"
        if not put_data.empty:
            put_data["Option_Type"] = "Put"

        return pd.concat([call_data, put_data], ignore_index=True)

    except Exception as e:
        print(f"⚠️ Error fetching data for '{stock_code}': {e}")
        return pd.DataFrame()


# ----------------------------
# Calculate PCR & Max Pain
# ----------------------------
def calculate_pcr_max_pain(data):
    """Calculate PCR (OI & Volume) and Max Pain."""
    data["open_interest"] = pd.to_numeric(data["open_interest"], errors="coerce").fillna(0)
    data["total_quantity_traded"] = pd.to_numeric(data["total_quantity_traded"], errors="coerce").fillna(0)

    total_put_oi = data[data["Option_Type"] == "Put"]["open_interest"].sum()
    total_call_oi = data[data["Option_Type"] == "Call"]["open_interest"].sum()
    pcr_oi = total_put_oi / total_call_oi if total_call_oi else 0

    total_put_volume = data[data["Option_Type"] == "Put"]["total_quantity_traded"].sum()
    total_call_volume = data[data["Option_Type"] == "Call"]["total_quantity_traded"].sum()
    pcr_volume = total_put_volume / total_call_volume if total_call_volume else 0

    # Max Pain Calculation
    losses = []
    for sp in data["strike_price"].unique():
        call_loss = sum(max(0, sp - strike) * oi for strike, oi in zip(
            data[data["Option_Type"] == "Call"]["strike_price"],
            data[data["Option_Type"] == "Call"]["open_interest"]
        ))
        put_loss = sum(max(0, strike - sp) * oi for strike, oi in zip(
            data[data["Option_Type"] == "Put"]["strike_price"],
            data[data["Option_Type"] == "Put"]["open_interest"]
        ))
        losses.append((sp, call_loss + put_loss))

    max_pain = min(losses, key=lambda x: x[1])[0] if losses else None
    return {
        "PCR OI": round(pcr_oi, 2),
        "PCR Volume": round(pcr_volume, 2),
        "Max Pain": max_pain
    }


# ----------------------------
# Append Results to Excel
# ----------------------------
def append_to_excel(output, file_path="data/Indicators_Log.xlsx", sheet_name="Sheet1"):
    """Append results to Excel log file."""
    try:
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
            if sheet_name not in writer.book.sheetnames:
                output.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                start_row = writer.book[sheet_name].max_row
                output.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=start_row)
        print(f"✅ Data appended to {file_path}")
    except FileNotFoundError:
        # Create new file if not exists
        output.to_excel(file_path, sheet_name=sheet_name, index=False)
        print(f"📄 New file created at {file_path}")
    except Exception as e:
        print(f"⚠️ Error appending to Excel: {e}")


# ----------------------------
# Visualization: Open Interest
# ----------------------------
def plot_open_interest(data, stock_code):
    """Plot Call vs Put Open Interest for given stock."""
    call_data = data[data["Option_Type"] == "Call"]
    put_data = data[data["Option_Type"] == "Put"]

    fig = go.Figure()
    fig.add_trace(go.Bar(x=call_data["strike_price"], y=call_data["open_interest"],
                         name="Call OI", marker_color="blue", width=15))
    fig.add_trace(go.Bar(x=put_data["strike_price"], y=put_data["open_interest"],
                         name="Put OI", marker_color="orange", width=15))
    fig.update_layout(
        title=f"Open Interest - {stock_code}",
        xaxis_title="Strike Price",
        yaxis_title="Open Interest",
        barmode="group",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5)
    )
    fig.show()


# ----------------------------
# Visualization: Total Loss Curve
# ----------------------------
def plot_total_loss(data, stock_code):
    """Plot Total Loss curve and highlight Max Pain strike."""
    strike_prices = data["strike_price"].unique()
    total_loss = []
    for sp in strike_prices:
        call_loss = sum(max(0, sp - strike) * oi for strike, oi in zip(
            data[data["Option_Type"] == "Call"]["strike_price"],
            data[data["Option_Type"] == "Call"]["open_interest"]
        ))
        put_loss = sum(max(0, strike - sp) * oi for strike, oi in zip(
            data[data["Option_Type"] == "Put"]["strike_price"],
            data[data["Option_Type"] == "Put"]["open_interest"]
        ))
        total_loss.append(call_loss + put_loss)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=strike_prices, y=total_loss, mode="lines+markers",
                             name="Total Loss", line=dict(color="red")))
    max_pain = strike_prices[total_loss.index(min(total_loss))]
    fig.add_trace(go.Scatter(x=[max_pain], y=[min(total_loss)], mode="markers+text",
                             name="Max Pain", text=["Max Pain"], textposition="top center",
                             marker=dict(size=10, color="purple")))
    fig.update_layout(
        title=f"Total Loss vs Strike Price - {stock_code}",
        xaxis_title="Strike Price",
        yaxis_title="Total Loss",
        template="plotly_white",
        legend=dict(orientation="h", yanchor="bottom", y=-0.25, xanchor="center", x=0.5)
    )
    fig.show()
    return max_pain


# ----------------------------
# Fetch Underlying LTP
# ----------------------------
def fetch_underlying_ltp(stock_code):
    """Fetch underlying spot/futures LTP from Breeze."""
    try:
        quote_data = breeze.get_quotes(
            stock_code=stock_code,
            exchange_code="nse",
            expiry_date="",
            product_type="",
            right="others",
            strike_price="0"
        )
        df_quote = pd.DataFrame(quote_data.get("Success", []))
        return df_quote.iloc[0].get("ltp") if not df_quote.empty else None
    except Exception as e:
        print("⚠️ Error fetching underlying LTP:", e)
        return None
