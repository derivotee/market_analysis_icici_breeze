"""
trend_analysis.py
---------------------------------
Functions for tracking PCR OI, PCR Volume, and Max Pain trends
across multiple expiries and timeframes using ICICI Breeze API.
"""

import os
import logging
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
from datetime import datetime

from scripts.breeze_client import get_breeze
from scripts.oi_analysis import fetch_option_chain_data, calculate_pcr_max_pain

# Initialize Breeze API (keys stored locally in breeze_client.py)
breeze = get_breeze()

# ============== DIRECTORIES ==============
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

DATA_FILE = os.path.join(PROJECT_ROOT, "data", "3MIndicators.xlsx")
LOG_DIR = os.path.join(PROJECT_ROOT, "logs")
OUTPUT_IMAGE_FOLDER = os.path.join(PROJECT_ROOT, "results")
LOG_FILE_PATH = os.path.join(LOG_DIR, "option_chain_trend.log")

# Ensure directories exist
os.makedirs(os.path.join(PROJECT_ROOT, "data"), exist_ok=True)
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(OUTPUT_IMAGE_FOLDER, exist_ok=True)

# ============== LOGGING CONFIGURATION ==============
logging.basicConfig(
    filename=LOG_FILE_PATH,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    filemode="a"
)
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console_handler.setFormatter(formatter)
logging.getLogger().addHandler(console_handler)


# ----------------------------
# Append Live Data to Excel
# ----------------------------
def append_live_data_to_sheet(stock_code, expiry, file_path=DATA_FILE):
    """Fetch live option chain data and append PCR/Max Pain to Excel by expiry sheet."""
    data = fetch_option_chain_data(stock_code, expiry)
    if data.empty:
        logging.warning(f"No data fetched for expiry {expiry}. Skipping...")
        return

    indicators = calculate_pcr_max_pain(data)
    now = datetime.now()
    new_row = [now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), expiry,
               indicators["PCR OI"], indicators["PCR Volume"], indicators["Max Pain"]]

    if not os.path.exists(file_path):
        wb = Workbook()
        wb.save(file_path)

    wb = load_workbook(file_path)
    if expiry not in wb.sheetnames:
        ws = wb.create_sheet(expiry)
        ws.append(["Date", "Time", "Expiry", "PCR OI", "PCR Volume", "Max Pain"])
    else:
        ws = wb[expiry]

    ws.append(new_row)
    wb.save(file_path)
    logging.info(f"Appended data for expiry {expiry} in {file_path}")


# ----------------------------
# Load trend data
# ----------------------------
def load_trend_data(file_path=DATA_FILE):
    """Load all expiry sheets from Excel file."""
    if not os.path.exists(file_path):
        logging.error("Trend log not found.")
        return {}
    xls = pd.ExcelFile(file_path)
    dfs = {}
    for sheet in xls.sheet_names:
        df = xls.parse(sheet)
        if "Date" in df.columns and not df.empty:
            df["Date"] = pd.to_datetime(df["Date"])
            dfs[sheet] = df
        else:
            logging.warning(f"Skipping sheet {sheet} - no data yet.")
    return dfs


# ----------------------------
# Plot trends from Excel
# ----------------------------
def plot_trends_from_excel(file_path=DATA_FILE):
    """Plot PCR OI, PCR Volume, and Max Pain trends from Excel log."""
    dfs = load_trend_data(file_path)
    if not dfs:
        return

    def plot_column(column_name, y_label, marker):
        plt.figure(figsize=(12, 6))
        for sheet, df in dfs.items():
            plt.plot(df["Date"], df[column_name], marker=marker, linestyle='-', label=f'{sheet}')
        plt.title(f"Trend of {column_name} Over Time")
        plt.xlabel("Date")
        plt.ylabel(y_label)
        plt.legend()
        plt.grid(True)

        img_path = os.path.join(OUTPUT_IMAGE_FOLDER, f"{column_name}_trend.png")
        plt.savefig(img_path)
        plt.show()
        logging.info(f"Saved {column_name} trend plot: {img_path}")

    plot_column("PCR OI", "PCR OI", "o")
    plot_column("PCR Volume", "PCR Volume", "s")
    plot_column("Max Pain", "Max Pain", "d")
